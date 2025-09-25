from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os

# Nombre del archivo Excel
filename = "colores.xlsx"

# Verificar si el archivo existe
if os.path.exists(filename):
    # Si existe, lo abrimos
    wb = load_workbook(filename)
    ws = wb.active
else:
    # Si no existe, lo creamos
    wb = Workbook()
    ws = wb.active
    ws.title = "Colores"

# Datos de ejemplo (color en RGB)
colores = [
    ("Rojo", (255, 0, 0)),
    ("Verde", (0, 255, 0)),
    ("Azul", (0, 0, 255)),
    ("Amarillo", (255, 255, 0)),
    ("Cian", (0, 255, 255))
]

# Escribimos los datos en el Excel
for nombre, (r, g, b) in colores:
    # Última fila disponible
    row = ws.max_row + 1

    # Guardar el nombre
    ws.cell(row=row, column=1, value=nombre)

    # Guardar el valor RGB como texto
    ws.cell(row=row, column=2, value=f"({r},{g},{b})")

    # Poner el color en una celda usando fill
    hex_color = f"{r:02X}{g:02X}{b:02X}"  # Convertimos a HEX (ej: FF0000)
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    ws.cell(row=row, column=3, value="").fill = fill

# Guardamos el archivo
wb.save(filename)
print(f"Datos guardados en {filename}")
