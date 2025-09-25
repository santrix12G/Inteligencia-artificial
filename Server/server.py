from flask import Flask, request, jsonify
import requests
import os
from datetime import datetime
import cv2
import numpy as np
import serial
# import serial.tools.list_ports
import time
# import subprocess
# import platform
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os
import socket

app = Flask(__name__) 


# Configuración Arduino salida
ARDUINO_OUT_PORT = "COM3"
ARDUINO_OUT_BAUDRATE = 9600
ARDUINO_OUT = None



entradas_verdes=[]
entradas_rojos=[]
entradas_azules=[]

color1=""
color2=""
color3=""


perceptrones=[]


#etapa de entrenamiento

def entrenar_perceptron(arreglo_entradas):
    print("Que color va entrenar")
    color=input()
    veces=10
    while veces:
        response = requests.get(IP_WEBCAM_URL, timeout=4)
        if response.status_code != 200:
            print("⚠️ No se pudo obtener la foto")
            return

        img_array = np.frombuffer(response.content, dtype=np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

        if img is None:
            print("⚠️ Error al decodificar imagen")
            return

        # Nombre con fecha/hora
        filename = datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"
        filepath = os.path.join(UPLOAD_FOLDER, filename)

        cv2.imwrite(filepath, img)

        rgb=procesar_color_imagen(img,0.4)

        llenar_excel(color,rgb)
        arreglo_entradas.append(rgb)
        veces-=1

    
    return color



        


def conectar_arduino(ARDUINO_PORT,ARDUINO_BAUDRATE):
    try:
        print(f"🔄 Intentando conectar a {ARDUINO_PORT}...")    
        # Intentar conexión
        arduino1 = serial.Serial(
            port=ARDUINO_PORT,
            baudrate=ARDUINO_BAUDRATE,
            bytesize=serial.EIGHTBITS,
            parity=serial.PARITY_NONE,
            stopbits=serial.STOPBITS_ONE,
            timeout=1,
            write_timeout=1
        )
        
        time.sleep(2)  # Esperar a que Arduino se reinicie
        print(f"✅ Arduino conectado en: {ARDUINO_PORT}")
        ARDUINO = arduino1
        
        return ARDUINO
            
    except serial.SerialException as e:
            print(f"❌ Error en {ARDUINO_PORT}: {e}")
            # continue
    except Exception as e:
            print(f"❌ Error inesperado en {ARDUINO_PORT}: {e}")
            # continue
    
    print("❌ No se pudo conectar a ningún puerto Arduino")
    return None

def enviar_a_arduino(datos, ARDUINO):
    try:
        # Convertir datos a string y enviar
        mensaje = f"{datos}\n"  # Añadir newline para que Arduino sepa que terminó el mensaje
        ARDUINO.write(mensaje.encode())
        print(f"📤 Enviado a Arduino: {mensaje.strip()}")
        return True
        
    except Exception as e:
        print(f"❌ Error enviando a Arduino: {e}")
        return False

def recibir_de_arduino(ARDUINO):
    try:
        if ARDUINO.in_waiting > 0:
            linea = ARDUINO.readline().decode('utf-8', errors='ignore').strip()
            print("📥 Recibido de Arduino: ", linea)
            return linea
        return None
    except Exception as e:
        print(f"❌ Error leyendo de Arduino: {e}")
        return None

    



#perceptron

def predict(rgb_values,color1,color2,color3):
    
    color=-1
    salida_verde=Perceptron1.activacion(rgb_values)
    salida_rojo=Perceptron2.activacion(rgb_values)
    

    if salida_verde:
        return color1
    elif salida_rojo:
        return color2
    else:
        return color3
    

        

class Perceptron:
    def __init__(self,pesos,bias):
        self.pesos=pesos
        self.bias=bias
    
    def activacion(self, x):
        z = np.dot(self.pesos, x)
        #return z
        return 1 if z + self.bias> 0 else 0

    def poner_color(self,color):
        self.color=color
    
    def entrenamiento(self,entradas,salidas,epocas,ts):
        for _ in range(epocas):
            j=0
            error_total=0
            for entrada in entradas:
                z=self.activacion(entrada)
                error = salidas[j] - z
                error_total+=error**2
                for i in range(len(self.pesos)):
                    self.pesos[i]+=(ts*entrada[i]*error)
                self.bias+= ts * error
                j+=1


#creacion de perceptrones
Perceptron1=Perceptron([-0.16321143 ,0.19252114,-0.10862646],[0.50256194])
Perceptron2=Perceptron([ 0.3241371 ,0.03054328,-0.50709092],[0.80362685])


    
def procesar_color_imagen(image_data, region_centro_porcentaje=0.3):
    try:
        if image_data is None:
            return {"error": "Imagen no válida"}
        
        height, width = image_data.shape[:2]
        
        # Calcular región central
        centro_x, centro_y = width // 2, height // 2
        tamaño_region = int(min(width, height) * region_centro_porcentaje)
        
        # Coordenadas del recuadro central
        x1 = max(0, centro_x - tamaño_region // 2)
        y1 = max(0, centro_y - tamaño_region // 2)
        x2 = min(width, centro_x + tamaño_region // 2)
        y2 = min(height, centro_y + tamaño_region // 2)
        
        # Extraer región central
        region_central = image_data[y1:y2, x1:x2]
        
        if region_central.size == 0:
            return {"error": "Región central demasiado pequeña"}
        

        # Convertir de BGR a RGB (OpenCV usa BGR por defecto)
        region_rgb = cv2.cvtColor(region_central, cv2.COLOR_BGR2RGB)
        
        # Calcular promedios
        promedio_rgb = np.mean(region_rgb, axis=(0, 1)).astype(int)
        return promedio_rgb
        
    except Exception as e:
        return {"error": f"Error procesando color: {str(e)}"}

# Función para llenar Excel
def llenar_excel(color_nombre, rgb):
    filename = "colores_nuevos.xlsx"
    # Verificar si el archivo existe
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Colores"
    r, g, b = rgb

    # Última fila disponible
    row = ws.max_row + 1
    # Guardar el nombre
    ws.cell(row=row, column=1, value=color_nombre)
    # Guardar el valor RGB como texto
    ws.cell(row=row, column=2, value=f"({r},{g},{b})")
    # Poner el color en una celda usando fill
    hex_color = f"{r:02X}{g:02X}{b:02X}"  # Convertimos a HEX (ej: FF0000)
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    ws.cell(row=row, column=3, value="").fill = fill
    wb.save(filename)


#configuracion de unity
import socket
import time

# Configuración del servidor
HOST = '172.27.21.133'  # Dirección IP local
PORT = 5053        # Puerto para la comunicación

def enviar_a_unity(data_to_send):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((HOST, PORT))
        s.listen()
        print("Servidor iniciado, esperando conexiones...")
        conn, addr = s.accept()
        with conn:
            print(f"Conectado a {addr}")
            try:
                # Envía los datos codificados
                conn.sendall(data_to_send.encode('utf-8'))
                print(f"Enviado: {data_to_send}")
                time.sleep(1) # Pausa de 1 segundo
            except (BrokenPipeError, ConnectionResetError):
                print("Conexión perdida con el cliente.")    



# Configuración de servidor
UPLOAD_FOLDER = 'capturas'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
IP_WEBCAM_URL = "http://172.27.18.138:8080/photo.jpg"

CAPTURE_INTERVAL = 0.2

def capture_and_save():
    global ARDUINO_OUT
    try:
        response = requests.get(IP_WEBCAM_URL, timeout=5)
        if response.status_code != 200:
            print("⚠️ No se pudo obtener la foto")
            return

        img_array = np.frombuffer(response.content, dtype=np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

        if img is None:
            print("⚠️ Error al decodificar imagen")
            return

        # Nombre con fecha/hora
        filename = datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"
        filepath = os.path.join(UPLOAD_FOLDER, filename)

        cv2.imwrite(filepath, img)

        rgb=procesar_color_imagen(img,0.4)
        print(predict(rgb,color1,color2,color3))
        print(rgb)
        # if ARDUINO_OUT is not None:
        #     if(color_predecido!="Azul"):
        #         enviar_a_arduino(color_predecido,ARDUINO_OUT)
            
        # try:
        #     if(color_predecido=="Rojo"):
        #         enviar_a_unity("0")
        #         print("informacion enviada a unity")
        #     elif(color_predecido=="Verde"):
        #         enviar_a_unity("1")
        #         print("informacion enviada a unity")
        #     else:
        #         enviar_a_unity("2")
        #         print("informacion enviada a unity")
        # except Exception as e:
        #     print(f"❌ Error enviando a Unity: {e}")
        # llenar_excel(predict(rgb),rgb)
            
        # print(f"🎨 Color RGB: {rgb}, Predicción: {prediction}")

        print(f"✅ Foto guardada: {filepath} ({img.shape[1]}x{img.shape[0]})")

    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    #configuracion de arduino
    ARDUINO_OUT = conectar_arduino(ARDUINO_OUT_PORT, ARDUINO_OUT_BAUDRATE)

    
    color1=entrenar_perceptron(entradas_verdes)
    color2=entrenar_perceptron(entradas_rojos)
    color3=entrenar_perceptron(entradas_azules)

    Perceptron1.poner_color(color1)
    Perceptron2.poner_color(color2)
    #entradas de datos y entrenamiento de los perceptrones
    salidas_verdes = np.full(len(entradas_verdes), 1)
    salidas_rojos=np.full(len(entradas_rojos), 1)

    #entradas_verdes para perceptron 1
    entradas_verdes = np.vstack((entradas_verdes, entradas_rojos))
    entradas_verdes = np.vstack((entradas_verdes, entradas_azules))
    salidas_verdes = np.hstack((salidas_verdes, np.zeros(len(entradas_rojos))))
    salidas_verdes = np.hstack((salidas_verdes, np.zeros(len(entradas_azules))))
    Perceptron1=Perceptron([-0.16321143 ,0.19252114,-0.10862646],[0.50256194])
    Perceptron1.entrenamiento(entradas_verdes,salidas_verdes,20,0.01)

    #entradas rojos para perceptron 2
    entradas_rojos = np.vstack((entradas_rojos, entradas_verdes))
    entradas_rojos = np.vstack((entradas_rojos, entradas_azules))
    salidas_rojos = np.hstack((salidas_rojos, np.zeros(len(entradas_verdes))))
    salidas_rojos = np.hstack((salidas_rojos, np.zeros(len(entradas_azules))))
    Perceptron2.entrenamiento(entradas_rojos,salidas_rojos,20,0.01)


    while True:
        capture_and_save()
        time.sleep(CAPTURE_INTERVAL)

